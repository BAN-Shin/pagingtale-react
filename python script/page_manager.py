import os
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


BG = "#f3f3f3"
FG = "#2f2f2f"
BTN_BG = "#4a4545"
BTN_FG = "#ffffff"
PANEL_BG = "#ffffff"
BORDER = "#6f6868"
SELECT_BG = "#d9d9d9"
NOTE_FG = "#555555"
INSERT_LINE = "#222222"

WINDOW_W = 1000
WINDOW_H = 790

LEFT_X = 48
RIGHT_X = 690
PANEL_Y = 236
PANEL_W = 260
PANEL_H = 460

CENTER_ADD_X = 355
CENTER_CTRL_LABEL_X = 560
CENTER_CTRL_ICON_X = 606

SMALL_BTN_Y = 706
OPEN_BTN_Y = 736
NOTE_Y = 674


@dataclass(frozen=True)
class PageItem:
    page: int
    title: str = ""

    @property
    def label(self) -> str:
        if self.title:
            return f"{self.page:03} {self.title}"
        return f"{self.page:03}"


class DragListbox(tk.Listbox):
    def __init__(self, master, move_callback, line_widget: tk.Frame, **kwargs):
        super().__init__(master, **kwargs)
        self.move_callback = move_callback
        self.line_widget = line_widget
        self.drag_indices: list[int] = []
        self.drop_index: Optional[int] = None

        self.bind("<ButtonPress-1>", self._on_press, add="+")
        self.bind("<B1-Motion>", self._on_motion, add="+")
        self.bind("<ButtonRelease-1>", self._on_release, add="+")
        self.bind("<Leave>", self._hide_insert_line, add="+")

    def _on_press(self, event):
        index = self.nearest(event.y)
        current = list(self.curselection())

        if index in current and current:
            self.drag_indices = current
        else:
            self.drag_indices = [index]
            self.selection_clear(0, tk.END)
            self.selection_set(index)

        self.drop_index = self._calc_drop_index(event.y)
        self._show_insert_line(event.y)

    def _on_motion(self, event):
        self.drop_index = self._calc_drop_index(event.y)
        self._show_insert_line(event.y)

    def _on_release(self, event):
        self._hide_insert_line()
        if self.drag_indices and self.drop_index is not None:
            self.move_callback(self.drag_indices, self.drop_index)
        self.drag_indices = []
        self.drop_index = None

    def _calc_drop_index(self, y: int) -> int:
        size = self.size()
        if size == 0:
            return 0

        nearest_index = self.nearest(y)
        bbox = self.bbox(nearest_index)

        if not bbox:
            return nearest_index

        _x, y0, _w, h = bbox
        mid = y0 + h / 2

        if y < mid:
            return nearest_index
        return nearest_index + 1

    def _show_insert_line(self, y: int):
        size = self.size()

        if size == 0:
            self.line_widget.place_forget()
            return

        drop_index = self._calc_drop_index(y)

        if drop_index <= 0:
            bbox = self.bbox(0)
            line_y = 2 if not bbox else max(2, bbox[1] + 1)
        elif drop_index >= size:
            bbox = self.bbox(size - 1)
            if bbox:
                line_y = bbox[1] + bbox[3] + 1
            else:
                line_y = self.winfo_height() - 3
        else:
            bbox = self.bbox(drop_index)
            line_y = 2 if not bbox else bbox[1] + 1

        self.line_widget.place(
            x=4,
            y=line_y,
            width=max(10, self.winfo_width() - 8),
            height=4
        )
        self.line_widget.lift()

    def _hide_insert_line(self, _event=None):
        self.line_widget.place_forget()


class PageManagerApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PagingTale ページ管理")
        self.root.geometry(f"{WINDOW_W}x{WINDOW_H}")
        self.root.resizable(False, False)
        self.root.configure(bg=BG)

        self.project_dir: Optional[Path] = None
        self.pages_dir: Optional[Path] = None
        self.index_file: Optional[Path] = None
        self.material_dir: Optional[Path] = None
        self.toc_file: Optional[Path] = None

        self.before_items: list[PageItem] = []
        self.after_items: list[PageItem] = []

        self._build_ui()
        self._refresh_lists()

    def _build_ui(self):
        tk.Label(
            self.root,
            text="PagingTale",
            font=("Yu Gothic UI", 34),
            bg=BG,
            fg=FG
        ).place(x=48, y=32)

        tk.Label(
            self.root,
            text="ページ管理",
            font=("Yu Gothic UI", 14),
            bg=BG,
            fg=FG
        ).place(x=50, y=116)

        tk.Label(
            self.root,
            text="before",
            font=("Yu Gothic UI", 28),
            bg=BG,
            fg=FG
        ).place(x=48, y=176)

        self.before_count_label = tk.Label(
            self.root,
            text="000 page",
            font=("Yu Gothic UI", 14),
            bg=BG,
            fg=FG
        )
        self.before_count_label.place(x=222, y=188)

        tk.Label(
            self.root,
            text="after",
            font=("Yu Gothic UI", 28),
            bg=BG,
            fg=FG
        ).place(x=682, y=176)

        self.after_count_label = tk.Label(
            self.root,
            text="000 page",
            font=("Yu Gothic UI", 14),
            bg=BG,
            fg=FG
        )
        self.after_count_label.place(x=856, y=188)

        self.before_frame = tk.Frame(
            self.root,
            bg=PANEL_BG,
            highlightbackground=BORDER,
            highlightthickness=1
        )
        self.before_frame.place(x=LEFT_X, y=PANEL_Y, width=PANEL_W, height=PANEL_H)

        self.before_list = tk.Listbox(
            self.before_frame,
            selectmode=tk.EXTENDED,
            font=("Yu Gothic UI", 15),
            bg=PANEL_BG,
            fg=FG,
            bd=0,
            highlightthickness=0,
            selectbackground=SELECT_BG,
            selectforeground=FG,
            activestyle="none",
            exportselection=False
        )
        self.before_list.place(x=10, y=10, width=218, height=408)

        self.before_yscroll = tk.Scrollbar(
            self.before_frame,
            orient="vertical",
            command=self.before_list.yview
        )
        self.before_yscroll.place(x=232, y=10, width=14, height=408)

        self.before_xscroll = tk.Scrollbar(
            self.before_frame,
            orient="horizontal",
            command=self.before_list.xview
        )
        self.before_xscroll.place(x=10, y=424, width=236, height=18)

        self.before_list.configure(
            yscrollcommand=self.before_yscroll.set,
            xscrollcommand=self.before_xscroll.set
        )

        self.after_frame = tk.Frame(
            self.root,
            bg=PANEL_BG,
            highlightbackground=BORDER,
            highlightthickness=1
        )
        self.after_frame.place(x=RIGHT_X, y=PANEL_Y, width=PANEL_W, height=PANEL_H)

        self.after_insert_line = tk.Frame(
            self.after_frame,
            bg=INSERT_LINE,
            bd=0,
            highlightthickness=0
        )

        self.after_list = DragListbox(
            self.after_frame,
            move_callback=self.reorder_after_by_drag,
            line_widget=self.after_insert_line,
            selectmode=tk.EXTENDED,
            font=("Yu Gothic UI", 15),
            bg=PANEL_BG,
            fg=FG,
            bd=0,
            highlightthickness=0,
            selectbackground=SELECT_BG,
            selectforeground=FG,
            activestyle="none",
            exportselection=False
        )
        self.after_list.place(x=10, y=10, width=218, height=408)

        self.after_yscroll = tk.Scrollbar(
            self.after_frame,
            orient="vertical",
            command=self.after_list.yview
        )
        self.after_yscroll.place(x=232, y=10, width=14, height=408)

        self.after_xscroll = tk.Scrollbar(
            self.after_frame,
            orient="horizontal",
            command=self.after_list.xview
        )
        self.after_xscroll.place(x=10, y=424, width=236, height=18)

        self.after_list.configure(
            yscrollcommand=self.after_yscroll.set,
            xscrollcommand=self.after_xscroll.set
        )

        self._make_icon_button("▶", CENTER_ADD_X, 286, self.add_selected_before, size=34)
        tk.Label(
            self.root,
            text="ADD",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_ADD_X + 5, y=356)

        self._make_icon_button("△", CENTER_CTRL_ICON_X, 284, self.move_top)
        tk.Label(
            self.root,
            text="TOP へ",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_CTRL_LABEL_X, y=302)

        self._make_icon_button("▲", CENTER_CTRL_ICON_X, 376, self.move_up)
        tk.Label(
            self.root,
            text="UP",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_CTRL_LABEL_X + 34, y=394)

        self._make_icon_button("✕", CENTER_CTRL_ICON_X, 468, self.delete_selected_after)
        tk.Label(
            self.root,
            text="DELETE",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_CTRL_LABEL_X - 4, y=486)

        self._make_icon_button("▼", CENTER_CTRL_ICON_X, 560, self.move_down)
        tk.Label(
            self.root,
            text="DOWN",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_CTRL_LABEL_X + 4, y=578)

        self._make_icon_button("▽", CENTER_CTRL_ICON_X, 632, self.move_bottom)
        tk.Label(
            self.root,
            text="BOTTOM へ",
            font=("Yu Gothic UI", 12),
            bg=BG,
            fg=FG
        ).place(x=CENTER_CTRL_LABEL_X - 40, y=650)

        self._make_small_button("すべて選択", LEFT_X, SMALL_BTN_Y, self.select_all_before)
        self._make_small_button("選択解除", LEFT_X + 124, SMALL_BTN_Y, self.clear_before_selection)
        self._make_small_button("すべて選択", RIGHT_X, SMALL_BTN_Y, self.select_all_after)
        self._make_small_button("選択解除", RIGHT_X + 124, SMALL_BTN_Y, self.clear_after_selection)

        self.open_btn = tk.Button(
            self.root,
            text="ページを開く",
            font=("Yu Gothic UI", 18),
            bg=BTN_BG,
            fg=BTN_FG,
            activebackground=BTN_BG,
            activeforeground=BTN_FG,
            relief="flat",
            bd=0,
            command=self.open_project,
            cursor="hand2"
        )
        self.open_btn.place(x=50, y=OPEN_BTN_Y, width=258, height=42)

        self.export_btn = tk.Button(
            self.root,
            text="ページ書出し",
            font=("Yu Gothic UI", 18),
            bg=BTN_BG,
            fg=BTN_FG,
            activebackground=BTN_BG,
            activeforeground=BTN_FG,
            relief="flat",
            bd=0,
            command=self.export_pages,
            cursor="hand2"
        )
        self.export_btn.place(x=684, y=OPEN_BTN_Y, width=258, height=42)

        note = (
            "Shift + クリックで範囲選択 / Ctrl + クリックで個別選択\n"
            "目次.xlsm を選ぶと、全ページを表示し、目次がないページは直前の目次タイトルを引き継ぎます"
        )
        tk.Label(
            self.root,
            text=note,
            justify="left",
            font=("Yu Gothic UI", 9),
            bg=BG,
            fg=NOTE_FG
        ).place(x=48, y=NOTE_Y)

    def _make_icon_button(self, text: str, x: int, y: int, command, size: int = 30):
        tk.Button(
            self.root,
            text=text,
            font=("Yu Gothic UI Symbol", size),
            bg=BG,
            fg=BTN_BG,
            activebackground=BG,
            activeforeground=BTN_BG,
            relief="flat",
            bd=0,
            command=command,
            cursor="hand2"
        ).place(x=x, y=y, width=78, height=64)

    def _make_small_button(self, text: str, x: int, y: int, command):
        tk.Button(
            self.root,
            text=text,
            font=("Yu Gothic UI", 9),
            bg="#dfdfdf",
            fg=FG,
            activebackground="#dfdfdf",
            activeforeground=FG,
            relief="flat",
            bd=1,
            command=command,
            cursor="hand2"
        ).place(x=x, y=y, width=112, height=26)

    def choose_html_file(self) -> Path:
        file_path = filedialog.askopenfilename(
            title="挿入基準にする pages\\mov_part_XXX.html を選択",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")]
        )
        if not file_path:
            raise RuntimeError("ファイルが選択されませんでした。")
        return Path(file_path)

    def choose_toc_file(self) -> Optional[Path]:
        file_path = filedialog.askopenfilename(
            title="目次.xlsm を選択（任意）",
            filetypes=[("Excel files", "*.xlsm *.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return None
        return Path(file_path)

    def open_project(self):
        try:
            html_path = self.choose_html_file()
        except Exception as e:
            messagebox.showwarning("ページ管理", str(e))
            return

        self.pages_dir = html_path.parent
        self.project_dir = self.pages_dir.parent
        self.index_file = self.project_dir / "index.html"
        self.material_dir = self.project_dir / "material"

        all_pages = self._load_pages(self.pages_dir)
        toc_map: dict[int, str] = {}

        toc_path = self.choose_toc_file()
        self.toc_file = toc_path

        if toc_path:
            try:
                toc_map = self._load_titles_from_excel(toc_path)
            except Exception as e:
                messagebox.showwarning("目次読み込み", f"目次の読み込みに失敗しました。\n{e}")
                toc_map = {}

        resolved_titles = self._build_page_titles_for_all_pages(all_pages, toc_map)
        self.before_items = [PageItem(page=p, title=resolved_titles.get(p, "")) for p in all_pages]
        self.after_items = []
        self._refresh_lists()

    def _load_pages(self, pages_dir: Path) -> list[int]:
        pages = []
        for f in sorted(pages_dir.glob("mov_part_*.html")):
            m = re.search(r"mov_part_(\d{3})\.html$", f.name, re.IGNORECASE)
            if m:
                pages.append(int(m.group(1)))
        if not pages:
            raise RuntimeError("pages フォルダ内に mov_part_XXX.html が見つかりませんでした。")
        return pages

    def _load_titles_from_excel(self, toc_path: Path) -> dict[int, str]:
        if load_workbook is None:
            raise RuntimeError("openpyxl が見つかりません。pip install openpyxl を実行してください。")

        wb = load_workbook(toc_path, data_only=True, read_only=True)
        if "目次" not in wb.sheetnames:
            raise RuntimeError("目次 シートが見つかりませんでした。")

        ws = wb["目次"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        header = [str(v).strip() if v is not None else "" for v in rows[0]]
        header_map = {name: idx for idx, name in enumerate(header)}

        if "page" not in header_map or "title" not in header_map:
            raise RuntimeError("目次 シートに page / title 列が必要です。")

        page_idx = header_map["page"]
        title_idx = header_map["title"]
        visible_idx = header_map.get("visible")

        titles: dict[int, str] = {}
        for row in rows[1:]:
            page_val = row[page_idx] if page_idx < len(row) else None
            title_val = row[title_idx] if title_idx < len(row) else None
            visible_val = row[visible_idx] if visible_idx is not None and visible_idx < len(row) else 1

            if page_val in (None, ""):
                continue
            if visible_val not in (1, "1", True):
                continue

            try:
                page_num = int(page_val)
            except Exception:
                continue

            titles[page_num] = "" if title_val is None else str(title_val).strip()

        return titles

    def _build_page_titles_for_all_pages(self, all_pages: list[int], toc_titles: dict[int, str]) -> dict[int, str]:
        if not toc_titles:
            return {p: "" for p in all_pages}

        resolved: dict[int, str] = {}
        current_title = ""
        toc_set = set(toc_titles.keys())

        for page in all_pages:
            if page in toc_set:
                current_title = toc_titles.get(page, "") or current_title
            resolved[page] = current_title

        return resolved

    def _refresh_lists(self):
        self.before_list.delete(0, tk.END)
        self.after_list.delete(0, tk.END)

        for item in self.before_items:
            self.before_list.insert(tk.END, item.label)
        for item in self.after_items:
            self.after_list.insert(tk.END, item.label)

        self.before_count_label.config(text=f"{len(self.before_items):03} page")
        self.after_count_label.config(text=f"{len(self.after_items):03} page")

    def select_all_before(self):
        self.before_list.selection_set(0, tk.END)

    def clear_before_selection(self):
        self.before_list.selection_clear(0, tk.END)

    def select_all_after(self):
        self.after_list.selection_set(0, tk.END)

    def clear_after_selection(self):
        self.after_list.selection_clear(0, tk.END)

    def add_selected_before(self):
        indices = list(self.before_list.curselection())
        if not indices:
            return

        selected = [self.before_items[i] for i in indices]
        self.after_items.extend(selected)
        self._refresh_lists()

        start = len(self.after_items) - len(selected)
        for i in range(start, len(self.after_items)):
            self.after_list.selection_set(i)

    def delete_selected_after(self):
        indices = set(self.after_list.curselection())
        if not indices:
            return

        self.after_items = [item for i, item in enumerate(self.after_items) if i not in indices]
        self._refresh_lists()

    def move_top(self):
        indices = list(self.after_list.curselection())
        if not indices:
            return

        moving_set = set(indices)
        moving = [self.after_items[i] for i in indices]
        remaining = [item for i, item in enumerate(self.after_items) if i not in moving_set]

        self.after_items = moving + remaining
        self._refresh_lists()

        for i in range(len(moving)):
            self.after_list.selection_set(i)

    def move_bottom(self):
        indices = list(self.after_list.curselection())
        if not indices:
            return

        moving_set = set(indices)
        moving = [self.after_items[i] for i in indices]
        remaining = [item for i, item in enumerate(self.after_items) if i not in moving_set]

        start = len(remaining)
        self.after_items = remaining + moving
        self._refresh_lists()

        for i in range(start, len(self.after_items)):
            self.after_list.selection_set(i)

    def move_up(self):
        indices = list(self.after_list.curselection())
        if not indices or indices[0] == 0:
            return

        items = self.after_items[:]
        for idx in indices:
            items[idx - 1], items[idx] = items[idx], items[idx - 1]

        self.after_items = items
        self._refresh_lists()

        for idx in [i - 1 for i in indices]:
            self.after_list.selection_set(idx)

    def move_down(self):
        indices = list(self.after_list.curselection())
        if not indices or indices[-1] >= len(self.after_items) - 1:
            return

        items = self.after_items[:]
        for idx in reversed(indices):
            items[idx + 1], items[idx] = items[idx], items[idx + 1]

        self.after_items = items
        self._refresh_lists()

        for idx in [i + 1 for i in indices]:
            self.after_list.selection_set(idx)

    def reorder_after_by_drag(self, from_indices: list[int], drop_index: int):
        if not from_indices:
            return

        from_indices = sorted(set(i for i in from_indices if 0 <= i < len(self.after_items)))
        if not from_indices:
            return

        moving = [self.after_items[i] for i in from_indices]
        moving_set = set(from_indices)
        remaining = [item for i, item in enumerate(self.after_items) if i not in moving_set]

        insert_at = max(0, min(drop_index, len(remaining)))
        self.after_items = remaining[:insert_at] + moving + remaining[insert_at:]
        self._refresh_lists()

        for i in range(insert_at, insert_at + len(moving)):
            self.after_list.selection_set(i)

    def export_pages(self):
        if self.project_dir is None or self.pages_dir is None:
            messagebox.showwarning("確認", "先にページを開いてください。")
            return

        if not self.after_items:
            messagebox.showwarning("確認", "after にページがありません。")
            return

        try:
            self._export_overwrite_with_backup()
        except Exception as e:
            messagebox.showerror("書き出しエラー", str(e))
            return

        messagebox.showinfo("ページ管理", "書き出しが完了しました。")

    def _export_overwrite_with_backup(self):
        assert self.project_dir is not None
        assert self.pages_dir is not None
        assert self.index_file is not None
        assert self.material_dir is not None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = self.project_dir.parent / f"{self.project_dir.name}_backup_{timestamp}"
        shutil.copytree(self.project_dir, backup_dir)

        temp_pages_dir = self.project_dir / "__pages_export_tmp__"
        temp_material_dir = self.project_dir / "__material_export_tmp__"

        if temp_pages_dir.exists():
            shutil.rmtree(temp_pages_dir)
        if temp_material_dir.exists():
            shutil.rmtree(temp_material_dir)

        temp_pages_dir.mkdir(parents=True, exist_ok=True)
        temp_material_dir.mkdir(parents=True, exist_ok=True)

        try:
            export_entries = []
            for new_page, item in enumerate(self.after_items, start=1):
                export_entries.append(
                    {
                        "old_page": item.page,
                        "new_page": new_page,
                        "title": item.title,
                    }
                )

            for entry in export_entries:
                old_page = entry["old_page"]
                new_page = entry["new_page"]

                src_html = self.pages_dir / f"mov_part_{old_page:03}.html"
                if src_html.exists():
                    text = src_html.read_text(encoding="utf-8")
                    text = self._replace_page_refs_in_text(text, old_page, new_page)
                    dst_html = temp_pages_dir / f"mov_part_{new_page:03}.html"
                    dst_html.write_text(text, encoding="utf-8")

                src_page_dir = self.material_dir / f"page_{old_page:03}"
                dst_page_dir = temp_material_dir / f"page_{new_page:03}"

                if src_page_dir.exists():
                    shutil.copytree(src_page_dir, dst_page_dir)
                    self._rename_and_rewrite_material(dst_page_dir, old_page, new_page)

            if self.index_file.exists():
                index_text = self.index_file.read_text(encoding="utf-8")
                index_text = re.sub(
                    r"(var\s+TOTAL_PAGES\s*=\s*)\d+",
                    rf"\g<1>{len(self.after_items)}",
                    index_text,
                )
                self.index_file.write_text(index_text, encoding="utf-8")

            if self.toc_file and self.toc_file.exists():
                self._rewrite_toc_excel(self.toc_file, export_entries)

            old_pages_backup = self.project_dir / f"__pages_old_{timestamp}__"
            if old_pages_backup.exists():
                shutil.rmtree(old_pages_backup, ignore_errors=True)
            shutil.move(str(self.pages_dir), str(old_pages_backup))
            shutil.move(str(temp_pages_dir), str(self.pages_dir))
            shutil.rmtree(old_pages_backup, ignore_errors=True)

            old_material_backup = self.project_dir / f"__material_old_{timestamp}__"
            if old_material_backup.exists():
                shutil.rmtree(old_material_backup, ignore_errors=True)
            shutil.move(str(self.material_dir), str(old_material_backup))
            shutil.move(str(temp_material_dir), str(self.material_dir))
            shutil.rmtree(old_material_backup, ignore_errors=True)

        except Exception:
            if temp_pages_dir.exists():
                shutil.rmtree(temp_pages_dir, ignore_errors=True)
            if temp_material_dir.exists():
                shutil.rmtree(temp_material_dir, ignore_errors=True)
            raise

    def _replace_page_refs_in_text(self, text: str, old_page: int, new_page: int) -> str:
        old3 = f"{old_page:03}"
        new3 = f"{new_page:03}"

        text = text.replace(f"page_{old3}", f"page_{new3}")
        text = re.sub(rf"(?<!\d){old3}(?!\d)", new3, text)
        text = re.sub(rf"(?<=_p){old3}(?=[^\d]|$)", new3, text)
        text = re.sub(rf"(?<=P){old3}(?=[^\d]|$)", new3, text)
        text = re.sub(rf"(?<=p){old3}(?=[^\d]|$)", new3, text)
        return text

    def _rename_and_rewrite_material(self, page_dir: Path, old_page: int, new_page: int):
        old3 = f"{old_page:03}"
        new3 = f"{new_page:03}"

        for root, _, files in os.walk(page_dir):
            root_path = Path(root)
            for name in sorted(files, key=len, reverse=True):
                old_path = root_path / name
                new_name = name.replace(f"_p{old3}", f"_p{new3}").replace(f"page_{old3}", f"page_{new3}")
                new_name = re.sub(rf"(?<!\d){old3}(?!\d)", new3, new_name)

                new_path = root_path / new_name
                if old_path != new_path:
                    old_path.rename(new_path)

        for file_path in page_dir.rglob("*"):
            if not file_path.is_file():
                continue

            suffix = file_path.suffix.lower()
            if suffix not in {".html", ".js", ".css", ".svg", ".txt", ".json", ".xml"}:
                continue

            try:
                text = file_path.read_text(encoding="utf-8")
            except Exception:
                continue

            text = self._replace_page_refs_in_text(text, old_page, new_page)
            file_path.write_text(text, encoding="utf-8")

    def _rewrite_toc_excel(self, toc_path: Path, export_entries: list[dict]):
        if load_workbook is None:
            raise RuntimeError("openpyxl が見つかりません。pip install openpyxl を実行してください。")

        backup_toc = toc_path.with_name(
            f"{toc_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{toc_path.suffix}"
        )
        shutil.copy2(toc_path, backup_toc)

        wb = load_workbook(toc_path, keep_vba=True)
        if "目次" not in wb.sheetnames:
            raise RuntimeError("目次.xlsm に『目次』シートが見つかりませんでした。")

        ws = wb["目次"]

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2:
            wb.save(toc_path)
            return

        headers = [ws.cell(row=1, column=col).value for col in range(1, max_col + 1)]
        header_map = {
            str(v).strip() if v is not None else "": idx + 1
            for idx, v in enumerate(headers)
        }

        if "page" not in header_map:
            raise RuntimeError("目次シートに page 列がありません。")

        page_col = header_map["page"]
        title_col = header_map.get("title")

        data_rows = []
        untouched_rows = []

        for row_idx in range(2, max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
            row_styles = [copy(ws.cell(row=row_idx, column=col)._style) for col in range(1, max_col + 1)]

            page_val = row_values[page_col - 1]
            old_page = self._safe_int(page_val)

            if old_page is None:
                untouched_rows.append((row_values, row_styles))
                continue

            data_rows.append(
                {
                    "row_values": row_values,
                    "row_styles": row_styles,
                    "old_page": old_page,
                }
            )

        row_map: dict[int, list[dict]] = {}
        for entry in data_rows:
            row_map.setdefault(entry["old_page"], []).append(entry)

        new_rows: list[tuple[list, list]] = []

        for entry in export_entries:
            old_page = entry["old_page"]
            new_page = entry["new_page"]
            new_title = entry["title"]

            candidates = row_map.get(old_page)
            if not candidates:
                continue

            source = candidates[0]
            new_values = list(source["row_values"])
            new_styles = [copy(style) for style in source["row_styles"]]

            new_values[page_col - 1] = new_page

            if title_col is not None and 1 <= title_col <= len(new_values):
                if new_title:
                    new_values[title_col - 1] = new_title

            new_rows.append((new_values, new_styles))

        new_rows.sort(key=lambda pair: self._safe_int(pair[0][page_col - 1]) or 10**9)
        final_rows = new_rows + untouched_rows

        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell._style = copy(ws.cell(row=1, column=col_idx)._style) if False else cell._style

        target_row = 2
        for row_values, row_styles in final_rows:
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=target_row, column=col_idx)
                cell.value = value
                if col_idx - 1 < len(row_styles):
                    cell._style = copy(row_styles[col_idx - 1])
            target_row += 1

        if target_row <= max_row:
            for row_idx in range(target_row, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None

        wb.save(toc_path)

    def _safe_int(self, value):
        try:
            if value is None or value == "":
                return None
            return int(value)
        except Exception:
            return None

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = PageManagerApp()
    app.run()